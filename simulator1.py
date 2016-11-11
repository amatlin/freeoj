import openpyxl
import numpy

wb = openpyxl.load_workbook('Data/MomPop/Results/MomPop2015.xlsm')
sheet = wb.get_sheet_by_name('FCOJ')

demandMatrix = numpy.zeros((7, 12))
for m in range(0, 12): # month
    c = m*8 + 4
    # NE
    for r in range(6, 19+1): # excel row
        demandMatrix[0][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # MA
    for r in range(20, 36+1): # excel row
        demandMatrix[1][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # SE
    for r in range(37, 48+1): # excel row
        demandMatrix[2][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # MW
    for r in range(49, 70+1): # excel row
        demandMatrix[3][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # DS
    for r in range(71, 86+1): # excel row
        demandMatrix[4][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # NW
    for r in range(87, 94+1): # excel row
        demandMatrix[5][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    # SW
    for r in range(95, 105+1): # excel row
        demandMatrix[6][m] = demandMatrix[0][m] + sheet.cell(row=r,column=c).value + sheet.cell(row=r,column=c+2).value + sheet.cell(row=r,column=c+4).value + sheet.cell(row=r,column=c+6).value 
    
# In September, we need to have 183 products for the North East
# Minimize cost of buying product + cost of shipping
# P07 Ilinois
# S14 Texas and S61 Ohio

# Buy OJ from 6 groves -> Ship to P07 -> Process -> Ship to S61
# Buy Future OJ from FLA -> Ship to P07 -> Process -> Ship to S61
# Buy Future FCOJ from FLA -> Ship to S61 -> Reconstitute





