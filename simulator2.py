import openpyxl
import numpy

StaticData = openpyxl.load_workbook('Data/StaticData.xlsx')
G_P = StaticData.get_sheet_by_name('G->PS')

MomPop2015 = openpyxl.load_workbook('Data/MomPop/Results/MomPop2015.xlsm')
FCOJ = MomPop2015.get_sheet_by_name('FCOJ')
grove = MomPop2015.get_sheet_by_name('grove')


demandMatrix = numpy.zeros((7, 12))
for m in range(0, 12): # month
    c = m*8 + 4
    # NE
    for r in range(6, 19+1): # excel row
        demandMatrix[0][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # MA
    for r in range(20, 36+1): # excel row
        demandMatrix[1][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # SE
    for r in range(37, 48+1): # excel row
        demandMatrix[2][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # MW
    for r in range(49, 70+1): # excel row
        demandMatrix[3][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # DS
    for r in range(71, 86+1): # excel row
        demandMatrix[4][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # NW
    for r in range(87, 94+1): # excel row
        demandMatrix[5][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    # SW
    for r in range(95, 105+1): # excel row
        demandMatrix[6][m] = demandMatrix[0][m] + FCOJ.cell(row=r,column=c).value + FCOJ.cell(row=r,column=c+2).value + FCOJ.cell(row=r,column=c+4).value + FCOJ.cell(row=r,column=c+6).value 
    
# In September, we need to have 183 products for the North East
# Minimize cost of buying product + cost of shipping
# P07 Ilinois
# S14 Texas and S61 Ohio

# Buy OJ from 6 groves -> Ship to P07 -> Process -> Ship to S61
# Buy Future OJ from FLA -> Ship to P07 -> Process -> Ship to S61
# Buy Future FCOJ from FLA -> Ship to S61 -> Reconstitute


groveMatrix = numpy.zeros((7, 12))
for r in range(0,5+1): # grove
    for c in range(0, 11+1): # month
        groveMatrix[r][c] = grove.cell(row=r+5,column=c+3).value / .0005

grove_to_processing = numpy.zeros((7, 12))
for c in range(0, 11+1): # month
    grove_to_processing[0][c] = G_P.cell(row=8,column=2).value * .22
    grove_to_processing[1][c] = G_P.cell(row=8,column=3).value * .22
    grove_to_processing[2][c] = G_P.cell(row=8,column=4).value * .22
    grove_to_processing[3][c] = G_P.cell(row=8,column=5).value * .22
    grove_to_processing[4][c] = G_P.cell(row=8,column=2).value * .22
    grove_to_processing[5][c] = G_P.cell(row=8,column=2).value * .22

manufacture_POJ = numpy.zeros((7,12))
manufacture_POJ.fill(2000)

processing_to_storage = numpy.zeros((7,12))

# have vectors which contain every different action at every step
# loop through all vectors with one function which determines price of those actions.
# put results into new vector

# there appears to be demand for each different type of product



